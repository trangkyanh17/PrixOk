#!/usr/bin/env python3
"""Side-effect-bounded V154 production smoke probe.

The probe loads live guard files directly by path rather than importing the bot
package, so it cannot create a Telegram client or a second AI worker. It tests
runtime dependencies plus the safety primitives that matter on the real phone.
"""

from __future__ import annotations

import argparse
import asyncio
import importlib.metadata
import importlib.util
import io
import json
import re
import sqlite3
import sys
import tempfile
import types
import zipfile
from pathlib import Path
from types import SimpleNamespace
from typing import Any

DEPENDENCIES = {
    "python-docx": "docx",
    "openpyxl": "openpyxl",
    "PyMuPDF": "pymupdf",
    "PyYAML": "yaml",
    "playwright": "playwright.async_api",
}


def _load(path: Path, name: str) -> Any:
    if not path.is_file():
        raise RuntimeError(f"probe target missing: {path}")
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load probe target: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def dependency_probe() -> dict[str, Any]:
    imported: dict[str, str] = {}
    for distribution, module_name in DEPENDENCIES.items():
        __import__(module_name)
        try:
            version = importlib.metadata.version(distribution)
        except importlib.metadata.PackageNotFoundError:
            version = "imported"
        imported[distribution] = version
    return {"dependencies": imported}


def _archive_probe(live_root: Path) -> dict[str, Any]:
    runtime = _load(
        live_root / "bot/modules/atri_attachment_runtime.py",
        "atri_v154_probe_attachment_runtime",
    )
    budget = runtime._ArchiveBudget()
    budget.reserve(1)

    with tempfile.TemporaryDirectory(prefix="atri-v154-archive-") as raw_tmp:
        tmp = Path(raw_tmp)
        inner_bytes = io.BytesIO()
        with zipfile.ZipFile(inner_bytes, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("proof.txt", b"atri-v154")
        outer = tmp / "outer.zip"
        with zipfile.ZipFile(outer, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("nested.zip", inner_bytes.getvalue())
        entries = runtime._expand_archive_safe(
            outer,
            tmp / "expanded",
            runtime._ArchiveBudget(),
        )
        names = {name for name, _path, _size in entries}
        if "nested.zip" not in names or "nested.zip!proof.txt" not in names:
            raise RuntimeError("nested ZIP expansion smoke failed")

        try:
            runtime._copy_archive_stream(
                io.BytesIO(b"12345"),
                tmp / "overflow.bin",
                expected=4,
            )
        except runtime.AttachmentRuntimeError as exc:
            if "ARCHIVE_MEMBER_STREAM_LIMIT" not in str(exc):
                raise
        else:
            raise RuntimeError("archive stream overrun was not blocked")
    return {"nested_zip": True, "stream_limit": True}


async def _audio_probe_async(system: Any) -> dict[str, Any]:
    class NeverDownload:
        calls = 0

        async def download(self, *, in_memory: bool = False):
            self.calls += 1
            raise RuntimeError("oversize audio should not download")

    media = SimpleNamespace(
        file_size=21 * 1024 * 1024,
        mime_type="audio/ogg",
        file_name="voice.ogg",
    )
    message = SimpleNamespace(voice=media, audio=None, document=None)
    selected = system._audio_media(message)
    if selected is None or selected[1] != "voice":
        raise RuntimeError("voice media classification failed")
    target = NeverDownload()
    result = await system._download_audio_inline(target, media, "voice")
    if result.get("audio_blocked") != "declared_size" or target.calls != 0:
        raise RuntimeError("oversize audio preflight did not block before download")
    return {"voice_classification": True, "oversize_preflight": True}


def _audio_and_tool_round_probe(live_root: Path) -> dict[str, Any]:
    system = _load(
        live_root / "bot/modules/atri_system_guard.py",
        "atri_v154_probe_system_guard",
    )
    audio = asyncio.run(_audio_probe_async(system))

    class FunctionCallResponse:
        def json(self) -> dict[str, Any]:
            return {
                "candidates": [
                    {
                        "content": {
                            "parts": [
                                {
                                    "functionCall": {
                                        "name": "probe",
                                        "args": {},
                                    }
                                }
                            ]
                        }
                    }
                ]
            }

    if not system._response_has_function_calls(FunctionCallResponse()):
        raise RuntimeError("tool-round functionCall detection failed")
    return {**audio, "tool_round_detection": True}


def _xlsx_probe(live_root: Path) -> dict[str, Any]:
    guard = _load(
        live_root / "bot/modules/atri_xlsx_formula_guard.py",
        "atri_v154_probe_xlsx_guard",
    )
    from openpyxl import Workbook, load_workbook

    raw = guard._safe_cell_value(lambda value: value, '=HYPERLINK("https://x")')
    formula = guard._safe_cell_value(lambda value: value, {"formula": "=SUM(A1:A2)"})
    if not isinstance(raw, str) or not raw.startswith("'="):
        raise RuntimeError("raw formula-like string was not escaped as text")
    if formula != "=SUM(A1:A2)":
        raise RuntimeError("explicit safe formula was not preserved")
    try:
        guard._safe_formula('=HYPERLINK("https://example.com","x")')
    except ValueError:
        pass
    else:
        raise RuntimeError("network formula was not rejected")

    for prefixed in (
        "=_xlfn.WEBSERVICE(A1)",
        '=_xlfn.RTD("prog.id",,"topic")',
        "=_xlws.UNKNOWNFUNC(A1)",
    ):
        try:
            guard._safe_formula(prefixed)
        except ValueError:
            pass
        else:
            raise RuntimeError(f"prefixed function bypass was not rejected: {prefixed}")

    with tempfile.TemporaryDirectory(prefix="atri-v154-xlsx-") as raw_tmp:
        path = Path(raw_tmp) / "probe.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = raw
        sheet["A2"] = formula
        workbook.save(path)
        reopened = load_workbook(path, data_only=False)
        try:
            if reopened.active["A1"].data_type != "s":
                raise RuntimeError("raw formula-like value reopened as non-string")
            if reopened.active["A2"].data_type != "f":
                raise RuntimeError("explicit formula did not reopen as formula")
        finally:
            reopened.close()
    return {
        "raw_text": True,
        "explicit_formula": True,
        "network_block": True,
        "prefixed_function_block": True,
    }


def _webapp_probe(live_root: Path) -> dict[str, Any]:
    guard = _load(
        live_root / "bot/modules/atri_webapp_safety_guard.py",
        "atri_v154_probe_webapp_guard",
    )
    blocked = False
    try:
        guard._validate_resolved_network_url("http://127.0.0.1:9229/json")
    except ValueError:
        blocked = True
    if not blocked:
        raise RuntimeError("loopback browser navigation was not blocked")
    public = "https://1.1.1.1/"
    if guard._validate_resolved_network_url(public) != public:
        raise RuntimeError("public literal URL validation failed")
    return {"loopback_block": True, "public_literal": True}


def _artifact_probe(live_root: Path) -> dict[str, Any]:
    guard = _load(
        live_root / "bot/modules/atri_artifact_relevance_guard.py",
        "atri_v154_probe_artifact_guard",
    )
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.executescript(
        """
        CREATE TABLE artifacts (
            id INTEGER PRIMARY KEY,
            artifact_ref TEXT,
            chat_key TEXT,
            message_id INTEGER,
            filename TEXT,
            active INTEGER,
            created_at INTEGER
        );
        CREATE TABLE chunks (
            id INTEGER PRIMARY KEY,
            artifact_id INTEGER,
            path TEXT,
            content TEXT
        );
        INSERT INTO artifacts(
            id, artifact_ref, chat_key, message_id, filename, active, created_at
        ) VALUES (1, 'artifact-probe', '100', 10, 'runtime.log', 1, 1);
        INSERT INTO chunks(id, artifact_id, path, content)
        VALUES (1, 1, 'runtime.log', 'quantum banana database timeout');
        """
    )

    class Index:
        @staticmethod
        def chat_key_from_message(message: Any) -> str:
            return str(message.chat.id)

        @staticmethod
        def _query_tokens(query: str) -> list[str]:
            return [
                token.casefold()
                for token in re.findall(r"[A-Za-z0-9_]{2,}", str(query))
            ][:24]

    def message(text: str, *, reply: Any = None) -> Any:
        return SimpleNamespace(
            chat=SimpleNamespace(id=100),
            text=text,
            caption="",
            reply_to_message=reply,
        )

    unrelated, reason = guard._is_relevant(
        Index, connection, message("weather hanoi today"), "weather hanoi today"
    )
    matched, matched_reason = guard._is_relevant(
        Index, connection, message("quantum banana"), "quantum banana"
    )
    followup, followup_reason = guard._is_relevant(
        Index, connection, message("sửa đi"), "sửa đi"
    )
    connection.execute("UPDATE artifacts SET active=0")
    inactive, inactive_reason = guard._is_relevant(
        Index, connection, message("sửa đi"), "sửa đi"
    )
    connection.close()
    if unrelated or reason != "unrelated":
        raise RuntimeError("unrelated artifact query was not rejected")
    if not matched or matched_reason != "chunk_match":
        raise RuntimeError("two-token artifact match was not accepted")
    if not followup or followup_reason != "short_followup":
        raise RuntimeError("short repair follow-up was not accepted")
    if inactive or inactive_reason != "no_artifact":
        raise RuntimeError("inactive artifact history was implicitly resurrected")
    return {
        "unrelated_block": True,
        "chunk_match": True,
        "short_followup": True,
        "inactive_history_block": True,
    }


def _sticker_probe(live_root: Path) -> dict[str, Any]:
    guard = _load(
        live_root / "bot/modules/atri_sticker_privacy_guard.py",
        "atri_v154_probe_sticker_guard",
    )
    saved = {
        name: sys.modules.get(name)
        for name in ("bot", "bot.modules", "bot.modules.atri_stickers")
    }
    with tempfile.TemporaryDirectory(prefix="atri-v154-sticker-") as raw_tmp:
        db_path = Path(raw_tmp) / "stickers.sqlite3"
        bot_pkg = types.ModuleType("bot")
        bot_pkg.__path__ = []  # type: ignore[attr-defined]
        modules_pkg = types.ModuleType("bot.modules")
        modules_pkg.__path__ = []  # type: ignore[attr-defined]
        stickers = types.ModuleType("bot.modules.atri_stickers")

        def connect() -> sqlite3.Connection:
            connection = sqlite3.connect(db_path)
            connection.row_factory = sqlite3.Row
            return connection

        def initialize() -> None:
            with connect() as connection:
                connection.execute(
                    """
                    CREATE TABLE IF NOT EXISTS stickers (
                        file_unique_id TEXT PRIMARY KEY,
                        first_chat_id INTEGER,
                        last_chat_id INTEGER,
                        last_seen_at INTEGER
                    )
                    """
                )

        def learn(**kwargs: Any) -> None:
            initialize()
            with connect() as connection:
                connection.execute(
                    """
                    INSERT INTO stickers(
                        file_unique_id, first_chat_id, last_chat_id, last_seen_at
                    ) VALUES (?, ?, ?, 1)
                    ON CONFLICT(file_unique_id) DO UPDATE SET
                        last_chat_id=excluded.last_chat_id,
                        last_seen_at=excluded.last_seen_at
                    """,
                    (
                        str(kwargs["file_unique_id"]),
                        int(kwargs["chat_id"]),
                        int(kwargs["chat_id"]),
                    ),
                )

        def candidates(chat_id: int, exclude_unique_id: str):
            del chat_id
            initialize()
            with connect() as connection:
                return connection.execute(
                    "SELECT * FROM stickers WHERE file_unique_id != ? ORDER BY file_unique_id",
                    (exclude_unique_id,),
                ).fetchall()

        def delete(file_unique_id: str) -> None:
            initialize()
            with connect() as connection:
                connection.execute(
                    "DELETE FROM stickers WHERE file_unique_id=?",
                    (file_unique_id,),
                )

        stickers._connect = connect  # type: ignore[attr-defined]
        stickers._initialize_sync = initialize  # type: ignore[attr-defined]
        stickers._learn_sync = learn  # type: ignore[attr-defined]
        stickers._candidate_rows_sync = candidates  # type: ignore[attr-defined]
        stickers._delete_sticker_sync = delete  # type: ignore[attr-defined]
        modules_pkg.atri_stickers = stickers  # type: ignore[attr-defined]
        bot_pkg.modules = modules_pkg  # type: ignore[attr-defined]
        sys.modules["bot"] = bot_pkg
        sys.modules["bot.modules"] = modules_pkg
        sys.modules["bot.modules.atri_stickers"] = stickers
        try:
            guard.install_atri_sticker_privacy_guard()
            stickers._learn_sync(file_unique_id="a", chat_id=100)
            stickers._learn_sync(file_unique_id="b", chat_id=200)
            rows_100 = stickers._candidate_rows_sync(100, "")
            rows_200 = stickers._candidate_rows_sync(200, "")
            ids_100 = {str(row["file_unique_id"]) for row in rows_100}
            ids_200 = {str(row["file_unique_id"]) for row in rows_200}
            if ids_100 != {"a"} or ids_200 != {"b"}:
                raise RuntimeError(
                    f"sticker chat scope failed chat100={ids_100} chat200={ids_200}"
                )
        finally:
            for name, previous in saved.items():
                if previous is None:
                    sys.modules.pop(name, None)
                else:
                    sys.modules[name] = previous
    return {"chat_scope": True}


def smoke_probe(live_root: Path) -> dict[str, Any]:
    result = {
        "archive": _archive_probe(live_root),
        "audio_tool_round": _audio_and_tool_round_probe(live_root),
        "artifact_rag": _artifact_probe(live_root),
        "sticker": _sticker_probe(live_root),
        "xlsx": _xlsx_probe(live_root),
        "webapp": _webapp_probe(live_root),
    }
    return result


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser()
    result.add_argument("action", choices=("deps", "smoke"))
    result.add_argument("--live-root", default="/app")
    return result


def main() -> int:
    args = parser().parse_args()
    try:
        if args.action == "deps":
            result = dependency_probe()
        else:
            result = smoke_probe(Path(args.live_root))
    except Exception as exc:
        print(
            json.dumps(
                {"ok": False, "error": type(exc).__name__ + ":" + str(exc)},
                ensure_ascii=False,
                sort_keys=True,
            )
        )
        return 1
    print(json.dumps({"ok": True, **result}, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
